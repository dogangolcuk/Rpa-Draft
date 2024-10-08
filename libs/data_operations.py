import pandas as pd
import os
from openpyxl import load_workbook
# from common_imports import pandas as pd, os
from data_store import DataStore
from libs.logging_utils import log_message
from openpyxl.utils import get_column_letter
from data_manipulator import DataManipulator
from libs.logging_utils import log_message
from config import plugins_dir



# Create an instance of DataManipulator
manipulator = DataManipulator()


def read_data_from_store(data_id):
    data_store = DataStore()
    return data_store.read_data(data_id)


# def manipulate_data(data):
#     if data is not None and not data.empty:
#         manipulated_data = data.copy()
#         manipulated_data["ENGLISH"] = manipulated_data["TURKISH"] + "deneme"
#         return manipulated_data
#     return None


def change_file_permissions(file_path):
    try:
        os.chmod(file_path, 0o666)
        print(f"Permissions changed for {file_path}")
    except OSError as e:
        print(f"Error changing permissions: {e}")


def write_to_excel_new_sheet(manipulated_data, excel_file, sheet_name='Manipulated_Data2'):
    try:
        # Create a new Pandas DataFrame from manipulated_data
        df = pd.DataFrame(manipulated_data)

        # Load the workbook
        book = load_workbook(excel_file)

        # Check if the specified sheet exists, if not, create a new one
        if sheet_name not in book.sheetnames:
            book.create_sheet(title=sheet_name)

        # Get the active sheet
        sheet = book[sheet_name]

        # Write the data to the specified sheet
        for idx, row in df.iterrows():
            for col_idx, value in enumerate(row):
                sheet[get_column_letter(col_idx + 1) + str(idx + 1)] = value

        # Save the workbook
        book.save(excel_file)

        print(f"Manipulated data saved to a new sheet '{
              sheet_name}' in {excel_file}")
    except Exception as e:
        print(f"Error occurred while writing to Excel: {e}")


def manipulate_and_write_data(cmd):
    data = read_data_from_store(cmd["getDataId"])
    print("Original Data")
    print(data)
    
    # Load the plugin
    pluginName = plugins_dir +"\\"+ cmd["usePLugin"]+ ".py"
    print(pluginName)
    # plugin = manipulator.load_plugin(pluginName)
    plugin = manipulator.load_plugin(pluginName)
    

    if plugin:
        try:
            # Manipulate data using the loaded plugin
            manipulated_data = manipulator.manipulate_data(data)
            if manipulated_data is not None:
                print("Manipulated Data")
                print(manipulated_data)

                change_file_permissions(cmd["filePath"])
                write_to_excel_new_sheet(manipulated_data, cmd["filePath"])
            else:
                log_message("Failed to manipulate data: Manipulated data is None")
        except Exception as e:
            log_message(f"Failed to manipulate data with plugin '{cmd['usePLugin']}': {e}")
            print("Error occurred during data manipulation.")
    else:
        log_message(f"Failed to load plugin '{cmd['usePLugin']}'")
        print(f"Failed to load plugin '{cmd['usePLugin']}'. Data manipulation aborted.")
